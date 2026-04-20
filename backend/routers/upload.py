import io
from database import get_db, log_change
from auth_utils import get_current_user
from typing import Optional
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends

router = APIRouter(prefix="/api", tags=["upload"])


@router.post("/upload-commissioning-data")
async def upload_commissioning_data(
    file: UploadFile = File(...),
    fiscalYear: str = Form("FY_25-26"),
    user: Optional[dict] = Depends(get_current_user)
):
    user_email = user.get("sub", "anonymous") if user else "anonymous"
    """Upload an Excel file with commissioning data and import to DB."""
    try:
        import openpyxl

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(400, "Empty workbook")

        # Read all rows as list of lists
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        # Find header row
        start_row = -1
        for i, row in enumerate(data):
            row_str = " ".join(str(c or "") for c in row)
            if "Project Name" in row_str and "Plan Actual" in row_str:
                start_row = i
                break

        if start_row == -1:
            raise HTTPException(400, "Could not find header row with 'Project Name' and 'Plan Actual'")

        headers = data[start_row]
        month_map = {
            "Apr": "apr", "May": "may", "Jun": "jun", "Jul": "jul",
            "Aug": "aug", "Sep": "sep", "Oct": "oct", "Nov": "nov",
            "Dec": "dec", "Jan": "jan", "Feb": "feb", "Mar": "mar",
        }

        # Find month column indices
        month_indices = {}
        for idx, h in enumerate(headers):
            h_str = str(h or "").strip()
            for prefix, db_col in month_map.items():
                if h_str.startswith(prefix):
                    month_indices[db_col] = idx
                    break

        success, failed = 0, 0
        current_project = {"name": "", "spv": ""}

        with get_db() as conn:
            for i in range(start_row + 1, len(data)):
                row = data[i]
                proj_name = str(row[1] or "").strip() if len(row) > 1 else ""
                if proj_name and proj_name not in ("", "nan"):
                    current_project["name"] = proj_name
                    current_project["spv"] = str(row[2] or "").strip() if len(row) > 2 else ""

                plan_actual_idx = next(
                    (idx for idx, h in enumerate(headers) if str(h or "").strip() == "Plan Actual"),
                    None,
                )
                if plan_actual_idx is None:
                    continue

                pa = str(row[plan_actual_idx] or "").strip() if len(row) > plan_actual_idx else ""
                if pa not in ("Plan", "Actual"):
                    continue

                if not current_project["name"]:
                    continue

                update_parts, params = [], []
                for db_col, col_idx in month_indices.items():
                    val = row[col_idx] if len(row) > col_idx else 0
                    try:
                        val = float(val) if val else 0
                    except (ValueError, TypeError):
                        val = 0
                    update_parts.append(f"{db_col} = %s")
                    params.append(val)

                if update_parts:
                    params.extend([fiscalYear, current_project["name"], pa])
                    conn.execute(
                        f"UPDATE commissioning_projects SET {', '.join(update_parts)} "
                        "WHERE fiscal_year = %s AND project_name = %s AND plan_actual = %s",
                        params,
                    )
                    success += 1
        
        log_change(user_email, "UPLOAD", "PROJECT_SHEET", fiscalYear, f"Updated {success} rows via Excel")
        return {
            "message": "Upload processed",
            "success_count": success,
            "failed_count": failed,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    fiscalYear: str = Form("FY_25-26"),
    user: Optional[dict] = Depends(get_current_user)
):
    user_email = user.get("sub", "anonymous") if user else "anonymous"
    """
    Full Excel import — parses workbook and imports projects.
    This is a simplified version; the TS excelParser logic is complex.
    We do a basic import here.
    """
    try:
        import openpyxl

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheets_found = wb.sheetnames
        imported = 0

        for sheet_name in sheets_found:
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            # Try to find header row
            start_row = -1
            for i, row in enumerate(data):
                row_str = " ".join(str(c or "") for c in row)
                if "Project Name" in row_str or "project_name" in row_str.lower():
                    start_row = i
                    break

            if start_row == -1:
                continue

            headers = [str(h or "").strip() for h in data[start_row]]

            # Map columns
            col_map = {}
            for idx, h in enumerate(headers):
                hl = h.lower().replace(" ", "_")
                if "project" in hl and "name" in hl:
                    col_map["project_name"] = idx
                elif "spv" in hl:
                    col_map["spv"] = idx
                elif "plan" in hl and "actual" in hl:
                    col_map["plan_actual"] = idx
                elif "category" in hl:
                    col_map["category"] = idx
                elif "priority" in hl:
                    col_map["priority"] = idx
                elif "charging" in hl:
                    col_map["charging_plan"] = idx
                elif "trail" in hl or "trial" in hl:
                    col_map["trial_run_plan"] = idx
                elif "cod" in hl:
                    col_map["cod_plan"] = idx
                elif "plot" in hl and "no" in hl:
                    col_map["plot_no"] = idx
                elif "capacity" in hl:
                    col_map["capacity"] = idx
                elif "type" in hl:
                    col_map["project_type"] = idx
                elif "location" in hl:
                    col_map["plot_location"] = idx

            if "project_name" not in col_map:
                continue

            with get_db() as conn:
                for i in range(start_row + 1, len(data)):
                    row = data[i]
                    pn = str(row[col_map.get("project_name", 1)] or "").strip() if len(row) > col_map.get("project_name", 1) else ""
                    if not pn or pn in ("nan", ""):
                        continue

                    pa = str(row[col_map.get("plan_actual", 3)] or "").strip() if col_map.get("plan_actual") and len(row) > col_map["plan_actual"] else "Plan"
                    cat = str(row[col_map.get("category", 0)] or "Solar").strip() if col_map.get("category") is not None and len(row) > col_map["category"] else "Solar"
                    spv = str(row[col_map.get("spv", 2)] or "").strip() if col_map.get("spv") and len(row) > col_map["spv"] else ""
                    
                    prio = str(row[col_map["priority"]] or "").strip() if "priority" in col_map and len(row) > col_map["priority"] else ""
                    cp = str(row[col_map["charging_plan"]] or "").strip() if "charging_plan" in col_map and len(row) > col_map["charging_plan"] else ""
                    tr = str(row[col_map["trial_run_plan"]] or "").strip() if "trial_run_plan" in col_map and len(row) > col_map["trial_run_plan"] else ""
                    cod = str(row[col_map["cod_plan"]] or "").strip() if "cod_plan" in col_map and len(row) > col_map["cod_plan"] else ""
                    plot_no = str(row[col_map["plot_no"]] or "").strip() if "plot_no" in col_map and len(row) > col_map["plot_no"] else ""
                    
                    try:
                        cap_raw = row[col_map["capacity"]] if "capacity" in col_map and len(row) > col_map["capacity"] else 0
                        if isinstance(cap_raw, str):
                            import re
                            cap_match = re.search(r"(\d+\.?\d*)", cap_raw)
                            cap = float(cap_match.group(1)) if cap_match else 0.0
                        else:
                            cap = float(cap_raw or 0)
                    except:
                        cap = 0.0

                    pt = str(row[col_map["project_type"]] or "PPA").strip() if "project_type" in col_map and len(row) > col_map["project_type"] else "PPA"
                    pl = str(row[col_map["plot_location"]] or "").strip() if "plot_location" in col_map and len(row) > col_map["plot_location"] else ""

                    conn.execute(
                        """INSERT INTO commissioning_projects
                        (fiscal_year, project_name, plan_actual, category, spv, project_type, plot_location, capacity, sno, 
                         priority, charging_plan, trial_run_plan, cod_plan, plot_no, is_deleted)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0)""",
                        (fiscalYear, pn, pa, cat, spv, pt, pl, cap, i, prio, cp, tr, cod, plot_no),
                    )
                    imported += 1

        log_change(user_email, "FULL_IMPORT", "PROJECT_SHEET", fiscalYear, f"Imported {imported} projects from Excel")
        return {
            "message": "Excel uploaded successfully",
            "projects_imported": imported,
            "sheets_found": sheets_found,
            "sheet_count": len(sheets_found),
        }

    except Exception as e:
        raise HTTPException(500, str(e))
